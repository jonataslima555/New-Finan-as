import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.common.by import By

# Carregue a planilha usando openpyxl
wb = openpyxl.load_workbook('investimento.xlsx')

# Configure o webdriver
webdriver_service = Service(GeckoDriverManager().install())
driver = webdriver.Firefox(service=webdriver_service)

# Selecione a planilha "Fibbs"
ws = wb['Fibbs']

# Inicie o contador de linhas em 2, pois a primeira linha contém os cabeçalhos
linha_atual = 2

for row in ws.iter_rows(min_row=2, values_only=True):
    acao = row[0]
    
    # Navegue até a página da ação
    driver.get(f'https://statusinvest.com.br/fundos-imobiliarios/{acao}/')
    #time.sleep(0.2)
    #botao = driver.find_element(By.XPATH, '//*[@id="hs-eu-confirmation-button"]')
    #botao.click()
    # Aguarde 5 segundos
    
    
    # Obtenha o preço atual e o último dividendo
    preco_atual_texto = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[1]/div[1]/div/div[1]/strong').text
    preco_atual = float(preco_atual_texto.replace('R$ ', '').replace(',', '.'))
    ultimo_dividendo_texto = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[7]/div[2]/div/div[1]/strong').text
    ultimo_dividendo = float(ultimo_dividendo_texto.replace('R$ ', '').replace(',', '.'))
    
    # Atualize a planilha
    ws.cell(row=linha_atual, column=2, value=preco_atual)
    ws.cell(row=linha_atual, column=3, value=ultimo_dividendo)
    print(f'\n_____________\n\n{acao} encontrada com sucesso\n_____________\n')
    
    # Incremente o contador de linhas
    linha_atual += 1

# Selecione a planilha "Ação"
ws = wb['Ação']

# Reinicie o contador de linhas para a nova planilha
linha_atual = 2

for row in ws.iter_rows(min_row=2, values_only=True):
    acao = row[0]
    
    # Navegue até a página da ação
    driver.get(f'https://statusinvest.com.br/acoes/{acao}')
    
    # Aguarde 2.5 segundos
    time.sleep(0.2)
    
    # Obtenha o preço atual e o último dividendo
    preco_atual = float(driver.find_element(By.XPATH, '/html/body/main/div[2]/div/div[1]/div/div[1]/div/div[1]/strong').text.replace(',', '.'))
    ultimo_dividendo_texto = driver.find_element(By.XPATH, '/html/body/main/div[2]/div/div[1]/div/div[4]/div/div[2]/div/span[2]').text
    ultimo_dividendo = float(ultimo_dividendo_texto.replace('R$ ', '').replace(',', '.'))
    
    # Atualize a planilha
    ws.cell(row=linha_atual, column=2, value=preco_atual)
    ws.cell(row=linha_atual, column=3, value=ultimo_dividendo)
    print(f'\n_____________\n\n{acao} encontrada com sucesso\n_____________\n')
    
    # Incremente o contador de linhas
    linha_atual += 1

# Salve a planilha atualizada
wb.save('investimento.xlsx')

driver.quit()

print('\n\nAções atualizadas...\n\n')
